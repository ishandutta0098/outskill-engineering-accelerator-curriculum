from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import html as html_lib

wb = load_workbook('assets/Engineering Accelerator Program - C7.xlsx')
ws = wb.active

merged_map = {}
merge_spans = {}

for merge in ws.merged_cells.ranges:
    min_row, min_col = merge.min_row, merge.min_col
    max_row, max_col = merge.max_row, merge.max_col
    rowspan = max_row - min_row + 1
    colspan = max_col - min_col + 1
    merge_spans[(min_row, min_col)] = (rowspan, colspan)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            if not (r == min_row and c == min_col):
                merged_map[(r, c)] = 'covered'

def get_fill_color(cell):
    try:
        fill = cell.fill
        if fill and fill.patternType and fill.patternType != 'none':
            color = fill.fgColor
            if color.type == 'rgb':
                return '#' + color.rgb[2:]
    except Exception:
        pass
    return None

def get_font_color(cell):
    try:
        color = cell.font.color
        if color and color.type == 'rgb':
            return '#' + color.rgb[2:]
    except Exception:
        pass
    return None

def format_value(val):
    if val is None:
        return ''
    s = str(val).strip()
    s = html_lib.escape(s)
    s = s.replace('\n', '<br>')
    while s.endswith('<br>'):
        s = s[:-4].rstrip()
    return s

MAX_ROW = 20
MAX_COL = 9

rows_html = []
for row in range(1, MAX_ROW + 1):
    cells_html = []
    for col in range(1, MAX_COL + 1):
        if merged_map.get((row, col)) == 'covered':
            continue

        cell = ws.cell(row=row, column=col)
        rowspan, colspan = merge_spans.get((row, col), (1, 1))

        styles = []

        bg = get_fill_color(cell)
        if bg:
            styles.append(f'background-color: {bg}')
        elif row == 1:
            styles.append('background-color: #3d1c52')

        fc = get_font_color(cell)
        if fc:
            styles.append(f'color: {fc}')
        elif row == 1:
            styles.append('color: #ffffff')

        try:
            if cell.font and cell.font.bold:
                styles.append('font-weight: bold')
        except Exception:
            pass

        try:
            align = cell.alignment
            if align:
                if align.horizontal in ('center', 'left', 'right'):
                    styles.append(f'text-align: {align.horizontal}')
                if align.vertical == 'center':
                    styles.append('vertical-align: middle')
                elif align.vertical == 'top':
                    styles.append('vertical-align: top')
        except Exception:
            pass

        attrs = []
        if rowspan > 1:
            attrs.append(f'rowspan="{rowspan}"')
        if colspan > 1:
            attrs.append(f'colspan="{colspan}"')
        if styles:
            attrs.append(f'style="{"; ".join(styles)}"')

        tag = 'th' if row == 1 else 'td'
        content = format_value(cell.value)
        attr_str = ' ' + ' '.join(attrs) if attrs else ''
        cells_html.append(f'<{tag}{attr_str}>{content}</{tag}>')

    rows_html.append(f'    <tr>{"".join(cells_html)}</tr>')

col_labels = ['#', 'Day', 'Time (IST)', 'Sprint', 'Sprint Title', 'Topics', 'Description', 'Outcomes', 'Tools']
col_widths = ['40px', '90px', '140px', '60px', '200px', '180px', '280px', '280px', '200px']

colgroup = '\n'.join(
    f'  <col style="width: {w}">' for w in col_widths
)

table_body = '\n'.join(rows_html)

html_output = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Engineering Accelerator Program - C7 Schedule</title>
  <style>
    body {{
      font-family: Arial, sans-serif;
      font-size: 13px;
      padding: 24px;
      background: #f9f9f9;
      color: #222;
    }}
    h1 {{
      font-size: 20px;
      margin-bottom: 16px;
      color: #3d1c52;
    }}
    .table-wrapper {{
      overflow-x: auto;
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      table-layout: fixed;
    }}
    th, td {{
      border: 1px solid #ccc;
      padding: 8px 10px;
      line-height: 1.5;
    }}
    th {{
      font-size: 13px;
      letter-spacing: 0.02em;
    }}
    td {{
      vertical-align: middle;
    }}
    br + br {{
      display: none;
    }}
  </style>
</head>
<body>
  <h1>Engineering Accelerator Program &mdash; C7 Schedule &amp; Roadmap</h1>
  <div class="table-wrapper">
    <table>
      <colgroup>
{colgroup}
      </colgroup>
      <tbody>
{table_body}
      </tbody>
    </table>
  </div>
</body>
</html>
"""

with open('assets/Engineering Accelerator Program - C7.html', 'w', encoding='utf-8') as f:
    f.write(html_output)

print("Done: assets/Engineering Accelerator Program - C7.html")
